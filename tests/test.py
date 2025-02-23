from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
import csv
from tests.conftest import ARCHIVE_FILE_PATH


def test_pdf_file(ARCHIVE_FILE_PATH):
    """Проверка PDF-файла"""
    with ZipFile(ARCHIVE_FILE_PATH, "r") as zipf:
        with zipf.open("pdf_file.pdf") as pdf_file:
            pdf_reader = PdfReader(pdf_file)

            # Проверка количества страниц
            assert len(pdf_reader.pages) == 5, "Должно быть 5 страниц"

            # Проверка текста на всех страницах
            text = "".join(page.extract_text() for page in pdf_reader.pages)
            assert "HR" in text


def test_csv_file(ARCHIVE_FILE_PATH):
    """Проверка CSV-файла"""
    with ZipFile(ARCHIVE_FILE_PATH, "r") as zipf:
        with zipf.open("csv_file.csv") as csv_file:
            csv_reader = csv.reader(csv_file.read().decode("utf-8").splitlines())
            csv_data = list(csv_reader)

            # Проверки структуры
            assert len(csv_data) == 78, "Должно быть 78 строк"
            assert all(
                len(row) == 2 for row in csv_data
            ), "Все строки должны иметь 2 колонки"

            # Поиск слова из файла
            found = any("Привет" in cell for row in csv_data for cell in row)
            assert found


def test_xlsx_file(ARCHIVE_FILE_PATH):
    """Проверка XLSX-файла"""
    with ZipFile(ARCHIVE_FILE_PATH, "r") as zipf:
        with zipf.open("xlsx_file.xlsx") as xlsx_file:
            wb = load_workbook(xlsx_file)
            sheet = wb.active

            # Проверки структуры
            assert sheet.max_row == 88, "Должно быть 88 строк"
            assert sheet.max_column == 2, "Должно быть 2 колонки"

            # Проверка n-й строки
            row_21 = str([cell.value for cell in sheet["21"]])
            assert "21" in row_21
